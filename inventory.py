import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_suppliers = {}
total_value_per_supplier = {}
product_inventory_under_10 = {}

#range：从2到最大行连续的数字list [2,3,4,5,6.....max_row+1],因为是不包含最后一个数字的，所以需要max_row+1
for prod_row in range(2, product_list.max_row + 1):
    #product_list.cell(row_num, column_num)可以定位到一个单元格，.value可以取出其中的内容
    supplier_name = product_list.cell(prod_row, 4).value
    inventory = product_list.cell(prod_row, 2).value
    price = product_list.cell(prod_row, 3).value
    product_number = product_list.cell(prod_row, 1).value
    inventory_price = product_list.cell(prod_row, 5)

    # calculation number of products per supplier
    if supplier_name in product_suppliers:
        current_num_prod = product_suppliers.get(supplier_name)
        product_suppliers[supplier_name] = current_num_prod + 1
    else:
        product_suppliers[supplier_name] = 1

    # calculation total value of inventories per supplier
    if supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calculation product number with less than 10 inventory
    if inventory < 10:
        product_inventory_under_10[int(product_number)] = inventory

    # add value for total inventory price
    inventory_price.value = inventory * price

print(product_suppliers)
print(total_value_per_supplier)
print(product_inventory_under_10)
inv_file.save("inventory_with_total_value.xlsx")